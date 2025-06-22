# ===============================================
# PRESSURE VESSEL COST CALCULATOR - MODERN UI
# Complete Application with AI Integration
# ===============================================

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import os
import sys
from pathlib import Path
import webbrowser
from datetime import datetime
import requests
import json
from typing import Dict, List, Tuple, Optional
import time

# Core calculator imports
import pypdf
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import io
from openai import OpenAI

# Set appearance mode and color theme
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class AIEnhancedPressureVesselCalculator:
    """AI-Enhanced Pressure Vessel Calculator with modern features"""
    def __init__(self, openai_api_key: Optional[str] = None, budget_mode: bool = True):
        self.openai_api_key = openai_api_key
        self.openai_client = None
        self.budget_mode = budget_mode
        self.model = "gpt-3.5-turbo" if budget_mode else "gpt-4"
        
        if openai_api_key:
            try:
                self.openai_client = OpenAI(api_key=openai_api_key)
            except Exception as e:
                print(f"Error initializing OpenAI: {e}")
        
        # Cost multipliers as specified
        self.cost_multipliers = {
            'heads': 9,    # $/lb - F&D Heads
            'shells': 13,  # $/lb - Cylindrical Shells
            'nozzles': 8,  # $/lb - Nozzles and fittings
            'flanges': 7,  # $/lb - Flanges
            'legs': 5,     # $/lb - Support legs
            'plates': 6    # $/lb - Reinforcing plates
        }

    def extract_pdf_text(self, pdf_path: str, page_numbers: List[int] = None) -> Dict[int, str]:
        """Extract text from specific PDF pages"""
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = pypdf.PdfReader(file)
                total_pages = len(pdf_reader.pages)
                
                extracted_text = {}
                
                # If no specific pages specified, extract key pages
                if page_numbers is None:
                    key_pages = [0, 2, 3, 4, 10, 11, 15, 17, 18, 19, 20, 22, 23]
                    page_numbers = [p for p in key_pages if p < total_pages]
                
                for page_num in page_numbers:
                    if page_num < total_pages:
                        try:
                            page = pdf_reader.pages[page_num]
                            text = page.extract_text()
                            extracted_text[page_num + 1] = text
                        except Exception as e:
                            print(f"Error extracting page {page_num + 1}: {e}")
                
                return extracted_text
        except Exception as e:
            raise Exception(f"Error reading PDF: {e}")

    def extract_vessel_info_traditional(self, pdf_text: Dict[int, str]) -> Dict:
        """Traditional regex extraction as fallback"""
        vessel_info = {}
        
        search_text = ""
        for page_num in [1, 2, 3]:
            if page_num in pdf_text:
                search_text += pdf_text[page_num] + "\n"
        
        # Extract vessel number
        vessel_patterns = [
            r'Vessel No[:\s]+([A-Z0-9-]+)',
            r'Tag Number[:\s]+([A-Z0-9-]+)',
            r'V-(\d+)',
            r'Vessel\s*#?\s*([A-Z0-9-]+)'
        ]
        
        for pattern in vessel_patterns:
            match = re.search(pattern, search_text, re.IGNORECASE)
            if match:
                vessel_info['vessel_number'] = match.group(1)
                break
        
        # Extract customer
        customer_patterns = [
            r'Customer[:\s]+([A-Za-z\s&,\.]+?)(?:\n|Contract|Designer|$)',
            r'Purchaser[:\s]+([A-Za-z\s&,\.]+?)(?:\n|Contract|Designer|$)'
        ]
        
        for pattern in customer_patterns:
            match = re.search(pattern, search_text, re.IGNORECASE)
            if match:
                vessel_info['customer'] = match.group(1).strip()
                break
        
        return vessel_info

    def extract_bill_of_materials(self, pdf_text: Dict[int, str]) -> Dict[str, List[Dict]]:
        """Extract bill of materials data from PDF content"""
        bom_data = {
            'heads': [],
            'shells': [],
            'legs': [],
            'nozzles': [],
            'flanges': [],
            'fasteners': [],
            'plates': []
        }
        
        # Look for BOM in pages 19-25
        bom_text = ""
        for page_num in range(19, 26):
            if page_num in pdf_text:
                bom_text += pdf_text[page_num] + "\n"
        
        # Extract Heads/Covers data
        heads_patterns = [
            r'H(\d+)\s+F&D Head\s+([A-Z0-9\s-]+)\s+([0-9.]+)\s*(?:\(min\.\))?\s+(\d+)\s+OD\s+([0-9.]+)\s+(\d+)',
        ]
        
        for pattern in heads_patterns:
            matches = re.findall(pattern, bom_text)
            for match in matches:
                try:
                    weight = float(match[4]) if len(match) > 4 else 0
                    quantity = int(match[5]) if len(match) > 5 else 1
                    
                    bom_data['heads'].append({
                        'item_id': f'H{match[0]}',
                        'type': 'F&D Head',
                        'material': match[1].strip(),
                        'thickness': match[2],
                        'weight': weight,
                        'quantity': quantity
                    })
                except (ValueError, IndexError):
                    continue
        
        return bom_data

    def extract_weight_summary(self, pdf_text: Dict[int, str]) -> Dict:
        """Extract weight summary data"""
        weight_summary = {}
        
        search_text = ""
        for page_num in range(12, 18):
            if page_num in pdf_text:
                search_text += pdf_text[page_num] + "\n"
        
        # Extract operating weight
        operating_patterns = [
            r'Operating Weight\s*\(lb\)\s*(\d+,?\d*)',
            r'Operating\s+(\d+,?\d*)',
        ]
        
        for pattern in operating_patterns:
            match = re.search(pattern, search_text)
            if match:
                weight_str = match.group(1).replace(',', '')
                try:
                    weight_summary['total_operating_weight'] = float(weight_str)
                    break
                except ValueError:
                    continue
        
        # Extract surface area
        surface_patterns = [
            r'Surface Area\s*\(ft[²2]\)\s*(\d+)',
            r'Surface Area\s*(\d+)',
        ]
        
        for pattern in surface_patterns:
            match = re.search(pattern, search_text)
            if match:
                try:
                    weight_summary['surface_area'] = float(match.group(1))
                    break
                except ValueError:
                    continue
        
        return weight_summary

    def calculate_costs(self, bom_data: Dict[str, List[Dict]]) -> Dict[str, Dict]:
        """Calculate costs based on extracted data and multipliers"""
        costs = {}
        
        # Calculate heads cost
        total_heads_weight = sum(item.get('weight', 0) * item.get('quantity', 1) for item in bom_data['heads'])
        if total_heads_weight > 0:
            costs['heads'] = {
                'weight': total_heads_weight,
                'rate': self.cost_multipliers['heads'],
                'total_cost': total_heads_weight * self.cost_multipliers['heads'],
                'description': f"{len(bom_data['heads'])} F&D Heads"
            }
        
        # Calculate shells cost
        total_shells_weight = sum(item.get('weight', 0) * item.get('quantity', 1) for item in bom_data['shells'])
        if total_shells_weight > 0:
            costs['shells'] = {
                'weight': total_shells_weight,
                'rate': self.cost_multipliers['shells'],
                'total_cost': total_shells_weight * self.cost_multipliers['shells'],
                'description': f"{len(bom_data['shells'])} Cylindrical Shells"
            }
        
        return costs

    def ai_comprehensive_analysis(self, pdf_text: Dict[int, str], traditional_vessel_info: Dict, 
                                 traditional_bom: Dict, weight_summary: Dict) -> Dict:
        """AI analysis with comprehensive cost estimation"""
        if not self.openai_client:
            return {'vessel_info': traditional_vessel_info, 'manual_costs': {}, 'validation': {}, 'analysis': ''}
        
        # Prepare condensed input
        search_text = ""
        for page_num in [1, 2, 19, 20]:
            if page_num in pdf_text:
                search_text += pdf_text[page_num][:800] + "\n"
        
        vessel_context = {
            'vessel_number': traditional_vessel_info.get('vessel_number', 'Unknown'),
            'customer': traditional_vessel_info.get('customer', 'Unknown'),
            'total_weight': weight_summary.get('total_operating_weight', 0),
            'surface_area': weight_summary.get('surface_area', 0),
        }
        
        prompt = f"""Analyze pressure vessel data and provide cost estimates.

PDF Sample: {search_text[:1000]}

Current Data: {json.dumps(vessel_context)}

Provide JSON response with exactly this structure:
{{
  "enhanced_vessel_info": {{
    "design_pressure": "50 psi",
    "material_grade": "SA-240 316",
    "design_temperature": "150°F"
  }},
  "manual_cost_estimates": {{
    "legs_fabrication": {{"unit_cost": 1500, "unit": "per leg", "total_cost": 6000}},
    "painting": {{"unit_cost": 12, "unit": "per ft²", "total_cost": 1896}},
    "testing_xray": {{"unit_cost": 75, "unit": "per ft", "total_cost": 3000}},
    "testing_ut": {{"unit_cost": 150, "unit": "per test", "total_cost": 900}},
    "transportation": {{"unit_cost": 2500, "unit": "per shipment", "total_cost": 2500}},
    "manway": {{"unit_cost": 1200, "unit": "each", "total_cost": 1200}}
  }},
  "validation": {{
    "overall_confidence": 8,
    "warnings": ["Check for missing flanges and fasteners"]
  }},
  "market_summary": "Current stainless steel 316 costs are stable."
}}

Use realistic 2024 US market rates. Return only valid JSON."""
        
        try:
            response = self.openai_client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=500,
                temperature=0.1
            )
            
            response_text = response.choices[0].message.content.strip()
            
            if response_text.startswith('```json'):
                response_text = response_text.replace('```json', '').replace('```', '').strip()
            
            result = json.loads(response_text)
            
            enhanced_vessel_info = traditional_vessel_info.copy()
            if 'enhanced_vessel_info' in result:
                enhanced_vessel_info.update(result['enhanced_vessel_info'])
            
            manual_costs = {}
            cost_estimates = result.get('manual_cost_estimates', {})
            surface_area = weight_summary.get('surface_area', 158)
            
            for service, cost_data in cost_estimates.items():
                if service == 'painting' and 'unit_cost' in cost_data:
                    actual_total = cost_data['unit_cost'] * surface_area
                else:
                    actual_total = cost_data.get('total_cost', 0)
                
                manual_costs[service] = {
                    'unit_cost': cost_data.get('unit_cost', 0),
                    'unit': cost_data.get('unit', 'each'),
                    'total_cost': actual_total,
                    'notes': f'AI estimate',
                    'source': 'AI estimate'
                }
            
            validation = result.get('validation', {'overall_confidence': 8})
            market_analysis = result.get('market_summary', 'Market analysis completed.')
            
            return {
                'vessel_info': enhanced_vessel_info,
                'manual_costs': manual_costs,
                'validation': validation,
                'analysis': market_analysis
            }
            
        except Exception as e:
            # Fallback with estimated costs
            surface_area = weight_summary.get('surface_area', 158)
            fallback_costs = {
                'legs_fabrication': {'unit_cost': 1500, 'unit': 'per leg', 'total_cost': 6000, 'notes': 'Fallback estimate', 'source': 'Fallback'},
                'painting': {'unit_cost': 12, 'unit': 'per ft²', 'total_cost': surface_area * 12, 'notes': 'Fallback estimate', 'source': 'Fallback'},
                'testing_xray': {'unit_cost': 75, 'unit': 'per ft', 'total_cost': 3000, 'notes': 'Fallback estimate', 'source': 'Fallback'},
                'transportation': {'unit_cost': 2500, 'unit': 'per shipment', 'total_cost': 2500, 'notes': 'Fallback estimate', 'source': 'Fallback'}
            }
            
            return {
                'vessel_info': traditional_vessel_info,
                'manual_costs': fallback_costs,
                'validation': {'overall_confidence': 7, 'warnings': ['AI analysis failed, using fallback estimates']},
                'analysis': f'Using fallback estimates: {str(e)}'
            }

    def create_excel_output(self, vessel_info: Dict, bom_data: Dict, 
                           weight_summary: Dict, costs: Dict,
                           manual_costs: Dict, ai_analysis: str,
                           validation: Dict, output_file: str) -> str:
        """Create Excel cost calculator"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cost Calculator"
        
        # Styles
        title_font = Font(bold=True, size=16, color="FFFFFF")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "AI-ENHANCED PRESSURE VESSEL COST CALCULATOR"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row = 3
        
        # Vessel Information
        vessel_info_items = [
            ("Vessel Number:", vessel_info.get('vessel_number', 'N/A')),
            ("Customer:", vessel_info.get('customer', 'N/A')),
            ("Design Pressure:", vessel_info.get('design_pressure', 'N/A')),
            ("Design Temperature:", vessel_info.get('design_temperature', 'N/A')),
            ("Material Grade:", vessel_info.get('material_grade', 'N/A')),
            ("Report Date:", datetime.now().strftime('%Y-%m-%d')),
            ("AI Confidence:", f"{validation.get('overall_confidence', 'N/A')}/10")
        ]
        
        for label, value in vessel_info_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Cost table headers
        headers = ['ITEM', 'DESCRIPTION', 'WEIGHT (lbs)', 'RATE ($/lb)', 'TOTAL COST ($)', 'SOURCE']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        
        # Material costs
        material_total = 0
        for component, cost_data in costs.items():
            ws.cell(row=row, column=1, value=component.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=cost_data['description'])
            ws.cell(row=row, column=3, value=f"{cost_data['weight']:.1f}")
            ws.cell(row=row, column=4, value=f"${cost_data['rate']}")
            ws.cell(row=row, column=5, value=f"${cost_data['total_cost']:,.2f}")
            ws.cell(row=row, column=6, value="BOM Extract")
            material_total += cost_data['total_cost']
            row += 1
        
        # Manual costs
        ai_total = 0
        for item_name, cost_data in manual_costs.items():
            ws.cell(row=row, column=1, value=item_name.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=cost_data.get('notes', ''))
            ws.cell(row=row, column=3, value="Service")
            ws.cell(row=row, column=4, value=f"${cost_data.get('unit_cost', 0)}/{cost_data.get('unit', 'each')}")
            ws.cell(row=row, column=5, value=f"${cost_data.get('total_cost', 0):,.2f}")
            ws.cell(row=row, column=6, value=cost_data.get('source', 'AI'))
            ai_total += cost_data.get('total_cost', 0)
            row += 1
        
        # Totals
        row += 1
        ws.cell(row=row, column=4, value="TOTAL PROJECT COST:")
        ws.cell(row=row, column=5, value=f"${material_total + ai_total:,.2f}")
        
        # Format column widths
        for col, width in enumerate([20, 35, 15, 15, 18, 15], 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        wb.save(output_file)
        return output_file

    def process_pdf(self, pdf_path: str, output_dir: str) -> str:
        """Main processing function"""
        # Extract text from PDF
        pdf_text = self.extract_pdf_text(pdf_path)
        
        if not pdf_text:
            raise Exception("Failed to extract text from PDF")
        
        # Traditional extractions
        vessel_info = self.extract_vessel_info_traditional(pdf_text)
        bom_data = self.extract_bill_of_materials(pdf_text)
        weight_summary = self.extract_weight_summary(pdf_text)
        costs = self.calculate_costs(bom_data)
        
        # AI analysis
        ai_results = self.ai_comprehensive_analysis(pdf_text, vessel_info, bom_data, weight_summary)
        
        final_vessel_info = ai_results.get('vessel_info', vessel_info)
        manual_costs = ai_results.get('manual_costs', {})
        validation = ai_results.get('validation', {'overall_confidence': 7})
        ai_analysis = ai_results.get('analysis', 'AI analysis not available')
        
        # Create output filename
        vessel_num = final_vessel_info.get('vessel_number', 'Unknown')
        output_filename = f"{vessel_num}_Cost_Calculator_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        # Create Excel output
        self.create_excel_output(
            final_vessel_info, bom_data, weight_summary, costs,
            manual_costs, ai_analysis, validation, output_path
        )
        
        return output_path

class ModernPressureVesselApp:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("AI-Enhanced Pressure Vessel Cost Calculator")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)
        
        # Variables
        self.pdf_path = tk.StringVar()
        self.openai_key = tk.StringVar()
        self.budget_mode = tk.BooleanVar(value=True)
        self.output_dir = tk.StringVar(value=str(Path.home() / "Documents"))
        self.email_address = tk.StringVar()
        self.last_output_path = None
        self.current_step = 1
        
        # Stats tracking
        self.files_processed = 0
        self.total_savings = 0
        self.success_rate = 100
        
        # Colors
        self.colors = {
            'primary': "#1f538d",
            'secondary': "#14375e", 
            'accent': "#00d4aa",
            'success': "#00c851",
            'warning': "#ffbb33",
            'error': "#ff4444",
            'light': "#f8f9fa",
            'dark': "#2c3e50"
        }
        
        # Load saved settings
        self.load_settings()
        
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the modern UI"""
        # Configure grid weights
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(0, weight=1)
        
        # Create sidebar
        self.create_sidebar()
        
        # Create main content area
        self.create_main_content()
        
        # Create status bar
        self.create_status_bar()
        
    def create_sidebar(self):
        """Create modern sidebar navigation"""
        sidebar = ctk.CTkFrame(self.root, width=280, corner_radius=0)
        sidebar.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        sidebar.grid_propagate(False)
        
        # App title and logo
        title_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=30)
        
        # App icon (you can replace with actual icon)
        icon_label = ctk.CTkLabel(
            title_frame, 
            text="⚙️", 
            font=ctk.CTkFont(size=32)
        )
        icon_label.pack(pady=(0, 10))
        
        app_title = ctk.CTkLabel(
            title_frame,
            text="Pressure Vessel\nCost Calculator",
            font=ctk.CTkFont(size=18, weight="bold"),
            justify="center"
        )
        app_title.pack()
        
        # Navigation buttons
        nav_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        nav_frame.pack(fill="x", padx=20, pady=20)
        
        self.nav_buttons = {}
        nav_items = [
            ("📁", "Upload PDF", "upload"),
            ("🤖", "AI Settings", "ai"),
            ("📊", "Generate Report", "generate"),
            ("📧", "Send Quote", "send"),
            ("⚙️", "Settings", "settings")
        ]
        
        for icon, text, key in nav_items:
            btn = ctk.CTkButton(
                nav_frame,
                text=f"{icon}  {text}",
                font=ctk.CTkFont(size=14),
                height=40,
                anchor="w",
                fg_color="transparent",
                text_color=("gray10", "gray90"),
                hover_color=("gray70", "gray30"),
                command=lambda k=key: self.switch_view(k)
            )
            btn.pack(fill="x", pady=2)
            self.nav_buttons[key] = btn
        
        # Quick stats
        stats_frame = ctk.CTkFrame(sidebar)
        stats_frame.pack(fill="x", padx=20, pady=20)
        
        ctk.CTkLabel(
            stats_frame,
            text="📈 Quick Stats",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(pady=(15, 10))
        
        self.stats_labels = {}
        stats = [
            ("Files Processed", str(self.files_processed)),
            ("Total Savings", f"${self.total_savings:,}"),
            ("Success Rate", f"{self.success_rate}%")
        ]
        
        for label, value in stats:
            stat_row = ctk.CTkFrame(stats_frame, fg_color="transparent")
            stat_row.pack(fill="x", padx=10, pady=2)
            
            ctk.CTkLabel(
                stat_row,
                text=label,
                font=ctk.CTkFont(size=11),
                text_color="gray"
            ).pack(side="left")
            
            stat_label = ctk.CTkLabel(
                stat_row,
                text=value,
                font=ctk.CTkFont(size=11, weight="bold")
            )
            stat_label.pack(side="right")
            self.stats_labels[label] = stat_label
        
        # Version info
        ctk.CTkLabel(
            sidebar,
            text="v1.0 - Powered by Refinery Connect",
            font=ctk.CTkFont(size=10),
            text_color="gray"
        ).pack(side="bottom", pady=20)
        
    def create_main_content(self):
        """Create main content area with different views"""
        self.main_frame = ctk.CTkFrame(self.root, corner_radius=0)
        self.main_frame.grid(row=0, column=1, sticky="nsew", padx=0, pady=0)
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(0, weight=1)
        
        # Create different view frames
        self.views = {}
        self.create_upload_view()
        self.create_ai_view()
        self.create_generate_view()
        self.create_send_view()
        self.create_settings_view()
        
        # Show upload view by default
        self.switch_view("upload")
        
    def create_upload_view(self):
        """Create PDF upload view"""
        view = ctk.CTkFrame(self.main_frame)
        self.views["upload"] = view
        
        # Header
        header = ctk.CTkFrame(view, height=80, fg_color="transparent")
        header.pack(fill="x", padx=30, pady=(30, 20))
        
        ctk.CTkLabel(
            header,
            text="📁 Upload PDF Document",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(side="left")
        
        # Step indicator
        step_label = ctk.CTkLabel(
            header,
            text="Step 1 of 3",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        step_label.pack(side="right")
        
        # Main upload area
        upload_frame = ctk.CTkFrame(view)
        upload_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
        
        # Drag & drop area (visual only)
        drop_area = ctk.CTkFrame(upload_frame, height=200, border_width=2, border_color="gray")
        drop_area.pack(fill="x", padx=40, pady=40)
        
        drop_content = ctk.CTkFrame(drop_area, fg_color="transparent")
        drop_content.pack(expand=True)
        
        ctk.CTkLabel(
            drop_content,
            text="📄",
            font=ctk.CTkFont(size=48)
        ).pack(pady=(20, 10))
        
        ctk.CTkLabel(
            drop_content,
            text="Drag & Drop PDF Here",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack()
        
        ctk.CTkLabel(
            drop_content,
            text="or click browse to select file",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        ).pack(pady=(0, 20))
        
        # File selection
        file_frame = ctk.CTkFrame(upload_frame, fg_color="transparent")
        file_frame.pack(fill="x", padx=40, pady=20)
        
        self.file_entry = ctk.CTkEntry(
            file_frame,
            textvariable=self.pdf_path,
            placeholder_text="No file selected...",
            height=40,
            font=ctk.CTkFont(size=12)
        )
        self.file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        browse_btn = ctk.CTkButton(
            file_frame,
            text="Browse Files",
            command=self.browse_file,
            height=40,
            width=120,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        browse_btn.pack(side="right")
        
        # File info (when file is selected)
        self.file_info_frame = ctk.CTkFrame(upload_frame)
        # Initially hidden
        
        # Next button
        next_btn = ctk.CTkButton(
            upload_frame,
            text="Next: Configure AI →",
            command=lambda: self.switch_view("ai"),
            height=45,
            width=200,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        next_btn.pack(side="bottom", anchor="e", padx=40, pady=20)
        
    def create_ai_view(self):
        """Create AI configuration view"""
        view = ctk.CTkFrame(self.main_frame)
        self.views["ai"] = view
        
        # Header
        header = ctk.CTkFrame(view, height=80, fg_color="transparent")
        header.pack(fill="x", padx=30, pady=(30, 20))
        
        ctk.CTkLabel(
            header,
            text="🤖 AI Configuration",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(side="left")
        
        step_label = ctk.CTkLabel(
            header,
            text="Step 2 of 3",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        step_label.pack(side="right")
        
        # AI settings
        settings_frame = ctk.CTkFrame(view)
        settings_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
        
        # API Key section
        api_section = ctk.CTkFrame(settings_frame)
        api_section.pack(fill="x", padx=30, pady=30)
        
        ctk.CTkLabel(
            api_section,
            text="🔑 OpenAI API Configuration",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", padx=20, pady=(20, 10))
        
        api_frame = ctk.CTkFrame(api_section, fg_color="transparent")
        api_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        ctk.CTkLabel(
            api_frame,
            text="API Key:",
            font=ctk.CTkFont(size=12)
        ).pack(anchor="w", pady=(0, 5))
        
        key_entry_frame = ctk.CTkFrame(api_frame, fg_color="transparent")
        key_entry_frame.pack(fill="x")
        
        self.api_key_entry = ctk.CTkEntry(
            key_entry_frame,
            textvariable=self.openai_key,
            show="*",
            placeholder_text="sk-...",
            height=35
        )
        self.api_key_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        get_key_btn = ctk.CTkButton(
            key_entry_frame,
            text="Get API Key",
            command=self.open_openai_link,
            height=35,
            width=100
        )
        get_key_btn.pack(side="right")
        
        # Test connection button
        self.test_btn = ctk.CTkButton(
            api_section,
            text="🔍 Test Connection",
            command=self.test_api_connection,
            height=35,
            width=150,
            fg_color="gray",
            hover_color="darkgray"
        )
        self.test_btn.pack(anchor="w", padx=20, pady=(0, 20))
        
        # Budget mode section
        budget_section = ctk.CTkFrame(settings_frame)
        budget_section.pack(fill="x", padx=30, pady=(0, 30))
        
        ctk.CTkLabel(
            budget_section,
            text="💰 Cost Settings",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", padx=20, pady=(20, 10))
        
        budget_frame = ctk.CTkFrame(budget_section, fg_color="transparent")
        budget_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        self.budget_checkbox = ctk.CTkCheckBox(
            budget_frame,
            text="Enable Budget Mode",
            variable=self.budget_mode,
            font=ctk.CTkFont(size=12)
        )
        self.budget_checkbox.pack(anchor="w")
        
        cost_info = ctk.CTkLabel(
            budget_frame,
            text="Budget Mode: ~$0.10 per analysis | Full Mode: ~$0.25 per analysis",
            font=ctk.CTkFont(size=10),
            text_color="gray"
        )
        cost_info.pack(anchor="w", pady=(5, 0))
        
        # Navigation buttons
        nav_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        nav_frame.pack(fill="x", padx=30, pady=20)
        
        back_btn = ctk.CTkButton(
            nav_frame,
            text="← Back",
            command=lambda: self.switch_view("upload"),
            height=45,
            width=120,
            fg_color="gray",
            hover_color="darkgray"
        )
        back_btn.pack(side="left")
        
        next_btn = ctk.CTkButton(
            nav_frame,
            text="Next: Generate Report →",
            command=lambda: self.switch_view("generate"),
            height=45,
            width=200,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        next_btn.pack(side="right")
        
    def create_generate_view(self):
        """Create report generation view"""
        view = ctk.CTkFrame(self.main_frame)
        self.views["generate"] = view
        
        # Header
        header = ctk.CTkFrame(view, height=80, fg_color="transparent")
        header.pack(fill="x", padx=30, pady=(30, 20))
        
        ctk.CTkLabel(
            header,
            text="📊 Generate Cost Report",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(side="left")
        
        step_label = ctk.CTkLabel(
            header,
            text="Step 3 of 3",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        step_label.pack(side="right")
        
        # Generation area
        gen_frame = ctk.CTkFrame(view)
        gen_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
        
        # Process button (prominent)
        process_section = ctk.CTkFrame(gen_frame)
        process_section.pack(fill="x", padx=30, pady=30)
        
        self.process_button = ctk.CTkButton(
            process_section,
            text="🚀 Generate Cost Calculator",
            command=self.process_pdf,
            height=60,
            font=ctk.CTkFont(size=18, weight="bold"),
            fg_color=self.colors['accent'],
            hover_color="#00b894"
        )
        self.process_button.pack(fill="x", padx=20, pady=20)
        
        # Progress section
        progress_section = ctk.CTkFrame(gen_frame)
        progress_section.pack(fill="x", padx=30, pady=(0, 30))
        
        ctk.CTkLabel(
            progress_section,
            text="📈 Processing Status",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", padx=20, pady=(20, 10))
        
        self.progress_bar = ctk.CTkProgressBar(progress_section, height=20)
        self.progress_bar.pack(fill="x", padx=20, pady=(0, 10))
        self.progress_bar.set(0)
        
        self.progress_label = ctk.CTkLabel(
            progress_section,
            text="Ready to process...",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        self.progress_label.pack(anchor="w", padx=20, pady=(0, 20))
        
        # Output directory
        output_section = ctk.CTkFrame(gen_frame)
        output_section.pack(fill="x", padx=30, pady=(0, 30))
        
        ctk.CTkLabel(
            output_section,
            text="📂 Output Location",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", padx=20, pady=(20, 10))
        
        output_frame = ctk.CTkFrame(output_section, fg_color="transparent")
        output_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        self.output_entry = ctk.CTkEntry(
            output_frame,
            textvariable=self.output_dir,
            height=35
        )
        self.output_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        browse_output_btn = ctk.CTkButton(
            output_frame,
            text="Browse",
            command=self.browse_output_dir,
            height=35,
            width=100
        )
        browse_output_btn.pack(side="right")
        
        # Results area
        self.results_frame = ctk.CTkFrame(gen_frame)
        self.results_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
        
        ctk.CTkLabel(
            self.results_frame,
            text="📋 Processing Log",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", padx=20, pady=(20, 10))
        
        self.status_text = ctk.CTkTextbox(
            self.results_frame,
            height=200,
            font=ctk.CTkFont(family="Consolas", size=11)
        )
        self.status_text.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
    def create_send_view(self):
        """Create send quote view"""
        view = ctk.CTkFrame(self.main_frame)
        self.views["send"] = view
        
        # Header
        header = ctk.CTkFrame(view, height=80, fg_color="transparent")
        header.pack(fill="x", padx=30, pady=(30, 20))
        
        ctk.CTkLabel(
            header,
            text="📧 Send Price Quote",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(side="left")
        
        # Send quote section
        send_frame = ctk.CTkFrame(view)
        send_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
        
        # Email section
        email_section = ctk.CTkFrame(send_frame)
        email_section.pack(fill="x", padx=30, pady=30)
        
        ctk.CTkLabel(
            email_section,
            text="✉️ Email Configuration",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", padx=20, pady=(20, 10))
        
        email_frame = ctk.CTkFrame(email_section, fg_color="transparent")
        email_frame.pack(fill="x", padx=20, pady=(0, 10))
        
        ctk.CTkLabel(
            email_frame,
            text="Recipient Email:",
            font=ctk.CTkFont(size=12)
        ).pack(anchor="w", pady=(0, 5))
        
        self.email_entry = ctk.CTkEntry(
            email_frame,
            textvariable=self.email_address,
            placeholder_text="customer@company.com",
            height=35
        )
        self.email_entry.pack(fill="x", pady=(0, 10))
        
        send_btn = ctk.CTkButton(
            email_section,
            text="📤 Send Price Quote",
            command=self.send_price_quote,
            height=45,
            width=200,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=self.colors['success'],
            hover_color="#00a847"
        )
        send_btn.pack(anchor="w", padx=20, pady=(0, 20))
        
        # Future integrations
        integrations_section = ctk.CTkFrame(send_frame)
        integrations_section.pack(fill="x", padx=30, pady=(0, 30))
        
        ctk.CTkLabel(
            integrations_section,
            text="🔗 Integrations",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", padx=20, pady=(20, 10))
        
        teamdesk_btn = ctk.CTkButton(
            integrations_section,
            text="📊 Send to TeamDesk (Coming Soon)",
            state="disabled",
            height=40,
            text_color="gray"
        )
        teamdesk_btn.pack(anchor="w", padx=20, pady=(0, 10))
        
        crm_btn = ctk.CTkButton(
            integrations_section,
            text="👥 Export to CRM (Coming Soon)",
            state="disabled",
            height=40,
            text_color="gray"
        )
        crm_btn.pack(anchor="w", padx=20, pady=(0, 20))
        
    def create_settings_view(self):
        """Create settings view"""
        view = ctk.CTkFrame(self.main_frame)
        self.views["settings"] = view
        
        # Header
        header = ctk.CTkFrame(view, height=80, fg_color="transparent")
        header.pack(fill="x", padx=30, pady=(30, 20))
        
        ctk.CTkLabel(
            header,
            text="⚙️ Application Settings",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(side="left")
        
        # Settings content
        settings_frame = ctk.CTkFrame(view)
        settings_frame.pack(fill="both", expand=True, padx=30, pady=(0, 30))
        
        # Theme settings
        theme_section = ctk.CTkFrame(settings_frame)
        theme_section.pack(fill="x", padx=30, pady=30)
        
        ctk.CTkLabel(
            theme_section,
            text="🎨 Appearance",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", padx=20, pady=(20, 10))
        
        theme_frame = ctk.CTkFrame(theme_section, fg_color="transparent")
        theme_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        ctk.CTkLabel(
            theme_frame,
            text="Theme:",
            font=ctk.CTkFont(size=12)
        ).pack(side="left", padx=(0, 10))
        
        theme_menu = ctk.CTkOptionMenu(
            theme_frame,
            values=["System", "Light", "Dark"],
            command=self.change_theme
        )
        theme_menu.pack(side="left")
        
        # About section
        about_section = ctk.CTkFrame(settings_frame)
        about_section.pack(fill="x", padx=30, pady=(0, 30))
        
        ctk.CTkLabel(
            about_section,
            text="ℹ️ About",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", padx=20, pady=(20, 10))
        
        about_text = """AI-Enhanced Pressure Vessel Cost Calculator v2.0

Modern, intelligent cost estimation for pressure vessel projects.
Built with Python, CustomTkinter, and OpenAI integration.

Features:
• PDF document analysis
• AI-powered cost estimation  
• Excel report generation
• Email integration
• Modern, intuitive interface"""
        
        ctk.CTkLabel(
            about_section,
            text=about_text,
            font=ctk.CTkFont(size=11),
            justify="left",
            text_color="gray"
        ).pack(anchor="w", padx=20, pady=(0, 20))
        
    def create_status_bar(self):
        """Create bottom status bar"""
        status_frame = ctk.CTkFrame(self.root, height=30, corner_radius=0)
        status_frame.grid(row=1, column=0, columnspan=2, sticky="ew")
        
        self.status_label = ctk.CTkLabel(
            status_frame,
            text="Ready",
            font=ctk.CTkFont(size=10),
            text_color="gray"
        )
        self.status_label.pack(side="left", padx=10, pady=5)
        
        # Connection status
        self.connection_label = ctk.CTkLabel(
            status_frame,
            text="● Disconnected",
            font=ctk.CTkFont(size=10),
            text_color="gray"
        )
        self.connection_label.pack(side="right", padx=10, pady=5)
        
    def switch_view(self, view_name):
        """Switch between different views"""
        # Hide all views
        for view in self.views.values():
            view.grid_remove()
        
        # Show selected view
        if view_name in self.views:
            self.views[view_name].grid(row=0, column=0, sticky="nsew")
        
        # Update navigation button states
        for key, btn in self.nav_buttons.items():
            if key == view_name:
                btn.configure(fg_color=("gray75", "gray25"))
            else:
                btn.configure(fg_color="transparent")
        
        # Update status bar if it exists
        if hasattr(self, 'status_label'):
            self.status_label.configure(text=f"Current view: {view_name.title()}")
        
    def browse_file(self):
        """Browse for PDF file"""
        filename = filedialog.askopenfilename(
            title="Select PDF File",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if filename:
            self.pdf_path.set(filename)
            self.update_file_info(filename)
    
    def update_file_info(self, filepath):
        """Update file information display"""
        if os.path.exists(filepath):
            size = os.path.getsize(filepath) / (1024 * 1024)  # MB
            message = f"✅ File selected: {os.path.basename(filepath)} ({size:.1f} MB)"
            
            # Only log if status system is ready
            if hasattr(self, 'status_text'):
                self.log_status(message)
            else:
                print(message)  # Fallback to console
    
    def browse_output_dir(self):
        """Browse for output directory"""
        directory = filedialog.askdirectory(title="Select Output Directory")
        if directory:
            self.output_dir.set(directory)
    
    def open_openai_link(self):
        """Open OpenAI API key page"""
        webbrowser.open("https://platform.openai.com/api-keys")
    
    def test_api_connection(self):
        """Test OpenAI API connection"""
        if not self.openai_key.get():
            messagebox.showwarning("Warning", "Please enter your OpenAI API key first.")
            return
        
        self.test_btn.configure(text="🔍 Testing...", state="disabled")
        self.connection_label.configure(text="● Testing...", text_color="orange")
        self.root.update()
        
        try:
            # Test the API key
            client = OpenAI(api_key=self.openai_key.get())
            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": "Hello"}],
                max_tokens=5
            )
            
            # If we get here, the API key works
            self.connection_label.configure(text="● Connected", text_color="green")
            self.test_btn.configure(text="✅ Connected", fg_color="green")
            messagebox.showinfo("Success", "API connection successful!")
            
        except Exception as e:
            self.connection_label.configure(text="● Error", text_color="red")
            self.test_btn.configure(text="❌ Failed", fg_color="red")
            messagebox.showerror("Connection Failed", f"API connection failed: {str(e)}")
        
        finally:
            self.root.after(3000, lambda: self.test_btn.configure(
                text="🔍 Test Connection", state="normal", fg_color="gray"
            ))
    
    def change_theme(self, theme):
        """Change application theme"""
        ctk.set_appearance_mode(theme)
    
    def log_status(self, message):
        """Add message to status text"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        
        # Only update status_text if it exists
        if hasattr(self, 'status_text'):
            self.status_text.insert("end", f"[{timestamp}] {message}\n")
            self.status_text.see("end")
            self.root.update()
        
        # Also update progress label if on generate view
        if hasattr(self, 'progress_label'):
            self.progress_label.configure(text=message)
        
        # Update status bar if it exists
        if hasattr(self, 'status_label'):
            self.status_label.configure(text=message)
    
    def update_stats(self):
        """Update the statistics display"""
        self.stats_labels["Files Processed"].configure(text=str(self.files_processed))
        self.stats_labels["Total Savings"].configure(text=f"${self.total_savings:,}")
        self.stats_labels["Success Rate"].configure(text=f"{self.success_rate}%")
    
    def process_pdf(self):
        """Process the PDF file"""
        if not self.pdf_path.get():
            messagebox.showerror("Error", "Please select a PDF file first")
            self.switch_view("upload")
            return
        
        if not os.path.exists(self.pdf_path.get()):
            messagebox.showerror("Error", "Selected PDF file does not exist")
            return
        
        if not os.path.exists(self.output_dir.get()):
            messagebox.showerror("Error", "Output directory does not exist")
            return
        
        # Disable process button
        self.process_button.configure(state="disabled", text="Processing...")
        self.progress_bar.set(0)
        self.status_text.delete("1.0", "end")
        
        # Save settings
        self.save_settings()
        
        # Start processing in background thread
        threading.Thread(target=self._process_pdf_thread, daemon=True).start()
    
    def _process_pdf_thread(self):
        """Background thread for PDF processing"""
        try:
            self.log_status("🚀 Starting PDF processing...")
            self.progress_bar.set(0.1)
            
            # Initialize calculator
            calculator = AIEnhancedPressureVesselCalculator(
                openai_api_key=self.openai_key.get() if self.openai_key.get() else None,
                budget_mode=self.budget_mode.get()
            )
            
            self.log_status("📖 Extracting text from PDF...")
            self.progress_bar.set(0.3)
            
            # Process PDF
            output_path = calculator.process_pdf(self.pdf_path.get(), self.output_dir.get())
            
            self.progress_bar.set(1.0)
            self.log_status(f"✅ SUCCESS! Excel file created: {os.path.basename(output_path)}")
            self.log_status(f"📂 Saved to: {output_path}")
            
            # Store path for later use
            self.last_output_path = output_path
            
            # Update stats
            self.files_processed += 1
            self.total_savings += 5000  # Example savings
            self.root.after(0, self.update_stats)
            
            # Show success dialog
            self.root.after(0, lambda: self._show_success_dialog(output_path))
            
        except Exception as e:
            self.progress_bar.set(0)
            self.log_status(f"❌ ERROR: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Processing Error", str(e)))
        
        finally:
            # Re-enable process button
            self.root.after(0, lambda: self.process_button.configure(
                state="normal", text="🚀 Generate Cost Calculator"
            ))
    
    def _show_success_dialog(self, output_path):
        """Show success dialog with options"""
        result = messagebox.askquestion(
            "Success!", 
            f"Cost calculator created successfully!\n\nFile: {os.path.basename(output_path)}\n\nWould you like to open the output folder?",
            icon='question'
        )
        
        if result == 'yes':
            self.open_output_folder(output_path)
        
        # Suggest next step
        if messagebox.askyesno("Next Step", "Would you like to send this quote via email?"):
            self.switch_view("send")
    
    def open_output_folder(self, file_path):
        """Open the output folder"""
        folder_path = os.path.dirname(file_path)
        
        if sys.platform == "win32":
            os.startfile(folder_path)
        elif sys.platform == "darwin":
            os.system(f"open '{folder_path}'")
        else:
            os.system(f"xdg-open '{folder_path}'")
    
    def send_price_quote(self):
        """Send the generated Excel sheet via email"""
        if not self.last_output_path or not os.path.exists(self.last_output_path):
            messagebox.showerror("Error", "No generated Excel file found. Please process a PDF first.")
            self.switch_view("generate")
            return

        email = self.email_address.get().strip()
        if not email:
            messagebox.showerror("Error", "Please enter the recipient email address.")
            return

        webhook_url = "https://automate.refineryconnect.com/webhook/4aea8e0f-07d4-4a06-ab02-83c196e36d21"

        try:
            if hasattr(self, 'status_text'):
                self.log_status(f"📤 Sending price quote to {email}...")
            
            with open(self.last_output_path, 'rb') as f:
                files = {
                    'file': (os.path.basename(self.last_output_path), f, 
                           'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                }
                data = {'email': email}
                response = requests.post(webhook_url, data=data, files=files, timeout=30)

            if response.status_code == 200:
                if hasattr(self, 'status_text'):
                    self.log_status("✅ Price quote sent successfully!")
                messagebox.showinfo("Success", "Price quote sent successfully!")
            else:
                if hasattr(self, 'status_text'):
                    self.log_status(f"❌ Failed to send price quote. Status: {response.status_code}")
                messagebox.showerror("Error", f"Failed to send price quote. Status code: {response.status_code}")
        except Exception as e:
            if hasattr(self, 'status_text'):
                self.log_status(f"❌ Error sending price quote: {e}")
            messagebox.showerror("Error", f"Error sending price quote: {e}")
    
    def save_settings(self):
        """Save user settings"""
        settings = {
            'openai_key': self.openai_key.get(),
            'budget_mode': self.budget_mode.get(),
            'output_dir': self.output_dir.get(),
            'email_address': self.email_address.get(),
            'files_processed': self.files_processed,
            'total_savings': self.total_savings,
            'success_rate': self.success_rate
        }
        
        try:
            settings_file = Path.home() / ".pressure_vessel_app_settings.json"
            with open(settings_file, 'w') as f:
                json.dump(settings, f)
        except:
            pass  # Ignore save errors
    
    def load_settings(self):
        """Load user settings"""
        try:
            settings_file = Path.home() / ".pressure_vessel_app_settings.json"
            if settings_file.exists():
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
                
                self.openai_key.set(settings.get('openai_key', ''))
                self.budget_mode.set(settings.get('budget_mode', True))
                self.output_dir.set(settings.get('output_dir', str(Path.home() / "Documents")))
                self.email_address.set(settings.get('email_address', ''))
                self.files_processed = settings.get('files_processed', 0)
                self.total_savings = settings.get('total_savings', 0)
                self.success_rate = settings.get('success_rate', 100)
        except:
            pass  # Ignore load errors
    
    def run(self):
        """Start the application"""
        self.root.mainloop()

def main():
    """Main entry point"""
    app = ModernPressureVesselApp()
    app.run()

if __name__ == "__main__":
    main()